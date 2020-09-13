#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import openpyxl 

path = input("Enter Path of excel file: ")

wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
m_row = sheet_obj.max_row 
folders = []


for i in range(1, m_row + 1): 
    a = sheet_obj.cell(row = i, column = 1)
    folders.append(a.value) 
    

for folder in folders: 
    os.makedirs(folder)


# In[ ]:




